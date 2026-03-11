"""
utils/matcher.py
================
All core logic for the AI Product Matching System.
Reusable by both streamlit_app.py and experiments.ipynb.

Usage in notebook:
    from utils.matcher import load_data, load_model, build_index, find_match

Usage in Streamlit:
    from utils.matcher import load_data, load_model, build_index, find_match

Run as terminal self-test:
    cd product_matching_system_Basemodel
    python utils/matcher.py
"""

import os
import re
import sys
import time
import logging
import warnings
import io
from datetime import datetime
from pathlib import Path

# ── Fix: Load .env and set HF_TOKEN BEFORE importing any HuggingFace library ─
# HuggingFace checks for the token at import time — if we don't set it first,
# the "unauthenticated requests" warning is already printed and can't be undone.
try:
    from dotenv import load_dotenv
    # Walk upward from this file's location to find .env in project root
    _env_search = Path(__file__).resolve()
    for _ in range(4):
        _env_candidate = _env_search / ".env"
        if _env_candidate.exists():
            load_dotenv(_env_candidate)
            break
        _env_search = _env_search.parent
except ImportError:
    pass  # python-dotenv not installed — env vars set manually are still used

# Set all three env var names HuggingFace uses across different versions
_hf_token = os.getenv("HF_TOKEN", "")
if _hf_token:
    os.environ["HF_TOKEN"]               = _hf_token
    os.environ["HUGGING_FACE_HUB_TOKEN"] = _hf_token  # older HF Hub versions
    os.environ["HUGGINGFACE_HUB_TOKEN"]  = _hf_token  # some transformers versions

# ── Suppress noisy warnings ───────────────────────────────────────────────────
os.environ["TOKENIZERS_PARALLELISM"]       = "false"
os.environ["HF_HUB_DISABLE_PROGRESS_BARS"] = "1"
logging.getLogger("sentence_transformers").setLevel(logging.ERROR)
logging.getLogger("transformers").setLevel(logging.ERROR)
logging.getLogger("huggingface_hub").setLevel(logging.ERROR)
warnings.filterwarnings("ignore")

# ── Heavy imports AFTER token is set ─────────────────────────────────────────
import numpy as np
import pandas as pd
import faiss
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sentence_transformers import SentenceTransformer


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

MATCH_THRESHOLD = 0.75

# ── Robust path resolution ─────────────────────────────────────────────────────
# Works whether this file is at:
#   project_root/utils/matcher.py   (normal)
#   project_root/matcher.py         (flat layout)
# Walks upward until it finds a sibling data/ folder.
def _find_data_dir() -> Path:
    """Walk upward from this file to find the data/ directory."""
    candidate = Path(__file__).resolve().parent
    for _ in range(4):
        data_dir = candidate / "data"
        if data_dir.is_dir():
            return data_dir
        candidate = candidate.parent
    # Last resort: data/ next to this file
    return Path(__file__).resolve().parent / "data"

DATA_DIR = _find_data_dir()
DEFAULT_DATA_PATH = DATA_DIR / "Matching List - Example 1 (1).xlsx"

# Results output file (written next to wherever the script is run from)
RESULTS_FILE = "product_matching_results.xlsx"

# Browser-like headers — helps bypass geo-restricted Arabic/Turkish e-commerce sites
SCRAPE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ar,en;q=0.9,tr;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# Site-specific CSS selectors for popular Saudi/regional e-commerce sites
SITE_SELECTORS = {
    "noon.com":        ["h1.product-title", "h1[data-testid='product-title']", "h1"],
    "jarir.com":       ["h1.page-title", "h1.product-name", "h1"],
    "amazon.sa":       ["#productTitle", "h1"],
    "amazon.com":      ["#productTitle", "h1"],
    "amazon.ae":       ["#productTitle", "h1"],
    "hepsiburada.com": ["h1.product-name", "h1[itemprop='name']", "h1"],
    "trendyol.com":    ["h1.pr-new-br", "h1.product-name", "h1"],
    "mediamarkt.com":  ["h1.product-name", "h1"],
    "extra.com":       ["h1.product-title", "h1"],
    "default":         ["h1[itemprop='name']", "h1.product-title", "h1.product-name", "h1"],
}

# Retailer shortcode → domain mapping
RETAILER_DOMAIN_MAP = {
    "TY":       "trendyol.com",
    "HB":       "hepsiburada.com",
    "NN_KSA":   "noon.com",
    "NN_UAE":   "noon.com",
    "AMZ_UAE":  "amazon.ae",
    "AMZ_KSA":  "amazon.sa",
    "AMZ_TR":   "amazon.com.tr",
    "MM_TR":    "mediamarkt.com.tr",
    "TKNS_TR":  "teknosa.com",
    "TKL_SA":   "takealot.com",
    "NHD":      "nahdionline.com",
    "GRTS":     "gratis.com",
    "C4_TR":    "carrefoursa.com",
    "C4_UAE":   "carrefouruae.com",
    "MSM":      "migros.com.tr",
    "A101":     "a101.com.tr",
    "ADW_KSA":  "al-dawaa.com",
    "CLK_SA":   "nahdi.sa",
    "EXTRA_SA": "extra.com",
    "JARIR":    "jarir.com",
}


# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING & ENRICHMENT
# ─────────────────────────────────────────────────────────────────────────────

def load_data(filepath=None) -> pd.DataFrame:
    """
    Load and enrich the product matching Excel dataset.

    Args:
        filepath: Path to the .xlsx file. Defaults to DEFAULT_DATA_PATH.
                  Always pass an explicit path from notebook/app for clarity.

    Returns:
        pd.DataFrame with these extra columns added:
            url_slug       — cleaned tokens extracted from the product URL
            enriched_title — title + unique slug tokens combined
            clean_title    — normalized text ready for embedding

    Raises:
        FileNotFoundError: if the filepath does not exist
        ValueError: if 'crawling title' column is missing from the file

    Example (notebook):
        df = load_data("../data/Matching List - Example 1 (1).xlsx")
        print(df.shape)
        df[["crawling title", "Link", "enriched_title"]].head()
    """
    if filepath is None:
        filepath = DEFAULT_DATA_PATH

    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(
            f"Data file not found: {filepath}\n"
            f"Check the filename and that it is inside the data/ folder."
        )

    df = pd.read_excel(filepath)

    # Validate required column exists
    if "crawling title" not in df.columns:
        raise ValueError(
            f"'crawling title' column not found in {filepath.name}.\n"
            f"Columns found: {list(df.columns)}"
        )

    # Keep only rows that have a title to index
    df = df.dropna(subset=["crawling title"])

    # Enrich: fill missing Link from ground-truth annotation column if present
    # 'Unnamed: 8' contains the human-verified correct link for mismatched rows
    correct_col = "Unnamed: 8"
    if correct_col in df.columns:
        mask_no_link = (
            df["Link"].isna() |
            df["Link"].astype(str).isin(["nan", "Page error", ""])
        )
        df.loc[mask_no_link, "Link"] = df.loc[mask_no_link, correct_col]

    # Drop rows still missing a valid link after enrichment
    df = df[df["Link"].astype(str).str.startswith("http")].reset_index(drop=True)

    if len(df) == 0:
        raise ValueError(
            f"No valid rows found in {filepath.name} after filtering. "
            f"Check that the 'Link' column contains http URLs."
        )

    # Add URL slug as extra matching signal
    # Product slugs often contain model numbers/specs not in the crawled title
    df["url_slug"]       = df["Link"].apply(_extract_slug_tokens)
    df["enriched_title"] = df.apply(_build_enriched_text, axis=1)
    df["clean_title"]    = df["enriched_title"].apply(clean_text)

    return df


def _extract_slug_tokens(url: str) -> str:
    """Extract meaningful product tokens from a URL slug."""
    if not isinstance(url, str) or not url.startswith("http"):
        return ""
    slug = url.rstrip("/").split("/")[-1].split("?")[0]
    slug = re.sub(r"[-_]", " ", slug)
    # Strip common noise patterns: long numeric IDs, platform-specific prefixes
    slug = re.sub(r"\b(p|dp|pd|pm|plid|hbc|hbcv|hbv|ref)\w*\b", "", slug, flags=re.I)
    slug = re.sub(r"\b\d{5,}\b", "", slug)
    slug = re.sub(r"\s+", " ", slug).strip()
    return slug


def _build_enriched_text(row) -> str:
    """Combine crawling title with unique URL slug tokens."""
    title_tokens = set(str(row["crawling title"]).lower().split())
    slug_tokens  = set(str(row["url_slug"]).lower().split())
    extra_tokens = slug_tokens - title_tokens  # only add tokens not already in title
    base = str(row["crawling title"])
    if extra_tokens:
        base = base + " " + " ".join(sorted(extra_tokens))
    return base


# ─────────────────────────────────────────────────────────────────────────────
# TEXT CLEANING
# ─────────────────────────────────────────────────────────────────────────────

def clean_text(text: str) -> str:
    """
    Normalize product text for embedding — Arabic, Latin, and digit safe.

    Steps:
        1. Lowercase
        2. Normalize storage units  (256gb → 256 gb)
        3. Keep Arabic (\u0600-\u06FF), Latin word chars, digits, spaces
        4. Collapse multiple spaces

    Example:
        clean_text("Samsung Galaxy 256GB - أزرق!")
        # → "samsung galaxy 256 gb أزرق"
    """
    text = str(text).lower()
    text = re.sub(r"(\d+)(gb|tb|mb)", r"\1 \2", text)
    text = re.sub(r"[^\w\s\u0600-\u06FF]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


# ─────────────────────────────────────────────────────────────────────────────
# MODEL LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_model(
    model_name: str = "paraphrase-multilingual-MiniLM-L12-v2",
    cache_folder: str = ".model_cache",
) -> SentenceTransformer:
    """
    Load the multilingual SentenceTransformer model.

    Supports Arabic, English, Turkish and 50+ other languages.
    First call downloads ~120MB and caches it in cache_folder/.
    Subsequent calls load from cache in ~3 seconds.

    Args:
        model_name:   HuggingFace model ID
        cache_folder: Local path to cache downloaded model weights

    Returns:
        SentenceTransformer instance ready for .encode()

    Note:
        HF_TOKEN must be set in .env BEFORE this function is called
        to avoid the "unauthenticated requests" warning. This file's
        top-level code handles that automatically on import.

    Example (notebook):
        model = load_model()
        emb = model.encode(["Samsung Galaxy S24"], normalize_embeddings=True)
        print(emb.shape)  # (1, 384)
    """
    import contextlib

    # Redirect stderr to suppress the C-level BertModel LOAD REPORT print
    # which fires at PyTorch level and cannot be caught by Python logging
    with open(os.devnull, "w") as devnull:
        old_stderr = sys.stderr
        sys.stderr  = devnull
        try:
            model = SentenceTransformer(model_name, cache_folder=cache_folder)
        finally:
            sys.stderr = old_stderr  # always restore even if load fails

    return model


# ─────────────────────────────────────────────────────────────────────────────
# FAISS INDEX
# ─────────────────────────────────────────────────────────────────────────────

def build_index(
    df: pd.DataFrame,
    model: SentenceTransformer,
    show_progress: bool = True,
) -> faiss.IndexFlatIP:
    """
    Build a FAISS inner-product index over all products in df.

    Uses cosine similarity via L2-normalised vectors + inner product,
    which gives identical results to cosine but runs faster in FAISS.

    Args:
        df:            DataFrame with a 'clean_title' column (from load_data)
        model:         Loaded SentenceTransformer (from load_model)
        show_progress: Show tqdm progress bar (True in notebook, False in app)

    Returns:
        faiss.IndexFlatIP — call index.search(query_vec, k) to retrieve top-k

    Example (notebook):
        index = build_index(df, model)
        print(f"Index has {index.ntotal} vectors, dim={index.d}")
    """
    if "clean_title" not in df.columns:
        raise ValueError(
            "'clean_title' column missing from df. "
            "Make sure you loaded the data with load_data() first."
        )

    embeddings = model.encode(
        df["clean_title"].tolist(),
        normalize_embeddings=True,
        show_progress_bar=show_progress,
    )
    embeddings = np.array(embeddings, dtype="float32")
    index      = faiss.IndexFlatIP(embeddings.shape[1])
    index.add(embeddings)
    return index


# ─────────────────────────────────────────────────────────────────────────────
# DOMAIN HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def get_domain(url_or_site: str) -> str:
    """
    Extract a bare domain from a URL, domain string, or www-prefixed domain.

    Example:
        get_domain("https://www.noon.com/product/123")  # → "noon.com"
        get_domain("www.trendyol.com")                  # → "trendyol.com"
    """
    s = str(url_or_site).strip().lower()
    s = re.sub(r"https?://(www\.)?", "", s)
    return s.split("/")[0]


def resolve_target_domain(raw_target: str) -> str:
    """
    Convert any target input format to a bare searchable domain string.

    Accepts:
        - Retailer codes  (TY       → trendyol.com)
        - www domains     (www.noon.com  → noon.com)
        - Full URLs       (https://extra.com/p/123 → extra.com)
        - Bare domains    (trendyol.com  → trendyol.com)
        - Empty / None    (""  → "")

    Example:
        resolve_target_domain("NN_KSA")                     # → "noon.com"
        resolve_target_domain("https://extra.com/product/") # → "extra.com"
    """
    if not raw_target or str(raw_target).strip() in ("", "nan", "N/A", "None"):
        return ""

    raw = str(raw_target).strip()

    # Retailer shortcode takes priority
    if raw.upper() in RETAILER_DOMAIN_MAP:
        return RETAILER_DOMAIN_MAP[raw.upper()]

    # Full URL
    if raw.startswith("http"):
        return get_domain(raw)

    # Plain domain (with or without www.)
    return re.sub(r"^www\.", "", raw.lower()).split("/")[0]


# ─────────────────────────────────────────────────────────────────────────────
# MATCHING ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def extract_model_numbers(text: str) -> set:
    """
    Extract product model codes and hardware specs from text.

    These are high-value exact-match signals that semantic similarity
    can miss when product names are in different languages.

    Catches: '2S9L1AA', 'NFK37211', 'T6644', '256GB', '330LT', '15W'

    Example:
        extract_model_numbers("HP 2S9L1AA 256GB laptop")
        # → {'2S9L1AA', '256GB'}
    """
    text  = str(text).upper()
    codes = re.findall(r"\b[A-Z0-9]{2,}[-/]?[A-Z0-9]{2,}\b", text)
    specs = re.findall(r"\b\d+\s?(?:GB|TB|MB|ML|GR|G|L|LT|W|MP)\b", text, re.I)
    return set(c.replace("-", "").replace("/", "") for c in codes + specs)


def find_match(
    query: str,
    model: SentenceTransformer,
    index: faiss.IndexFlatIP,
    df: pd.DataFrame,
    threshold: float = MATCH_THRESHOLD,
    target_filter: str = None,
) -> pd.DataFrame:
    """
    3-Signal Matching Engine — returns the top 5 most similar products.

    Scoring formula:
        combined = semantic * 0.60
                 + keyword_overlap * 0.20
                 + model_number_boost (capped at 0.20)

    For partial queries (<=4 words), keyword_overlap gets an extra 0.10 weight.

    Args:
        query:         Product title to search for (any language)
        model:         Loaded SentenceTransformer
        index:         FAISS index built from df (from build_index)
        df:            Reference product DataFrame (from load_data)
        threshold:     Minimum combined score to mark as "Match Found"
        target_filter: Domain, full URL, or retailer code to restrict
                       results to one specific site (optional)

    Returns:
        pd.DataFrame of top-5 results with columns:
            crawling title, Link, Similarity Score, Match Status
            + all original df columns
        DataFrame.attrs contains:
            domain_filtered (bool) — whether filter was applied
            domain_used     (str)  — which domain was used as filter

    Example (notebook):
        results = find_match(
            "Samsung Galaxy S24 256GB",
            model, index, df,
            threshold=0.70,
            target_filter="NN_KSA"
        )
        results[["crawling title", "Similarity Score", "Match Status", "Link"]]
    """
    if not query or not str(query).strip():
        raise ValueError("query cannot be empty")

    query_clean  = clean_text(query)
    query_tokens = set(query_clean.split())
    query_models = extract_model_numbers(query)
    is_partial   = len(query_tokens) <= 4

    query_embedding = model.encode(
        [query_clean], normalize_embeddings=True
    ).astype("float32")

    # ── Domain filter ─────────────────────────────────────────────────────────
    search_df     = df.copy()
    domain_filter = None

    if target_filter and str(target_filter).strip() not in ("", "N/A", "nan", "None"):
        # Resolve any format (code / URL / domain) to a bare domain
        resolved = resolve_target_domain(target_filter)
        if resolved:
            filtered = df[df["Link"].str.contains(resolved, case=False, na=False)]
            if len(filtered) >= 1:
                search_df     = filtered.reset_index(drop=True)
                domain_filter = resolved
            # If no products found for that domain, silently fall back to all products

    # ── FAISS vector search ───────────────────────────────────────────────────
    if domain_filter is not None and len(search_df) < len(df):
        # Build a temporary sub-index for the filtered subset only
        sub_emb = model.encode(
            search_df["clean_title"].tolist(),
            normalize_embeddings=True,
            show_progress_bar=False,
        ).astype("float32")
        sub_idx = faiss.IndexFlatIP(sub_emb.shape[1])
        sub_idx.add(sub_emb)
        top_k        = min(20, len(search_df))
        scores, ids  = sub_idx.search(query_embedding, top_k)
        results      = search_df.iloc[ids[0]].copy()
    else:
        top_k        = min(20, len(df))
        scores, ids  = index.search(query_embedding, top_k)
        results      = df.iloc[ids[0]].copy()

    results = results.copy()  # avoid SettingWithCopyWarning
    results["_semantic"] = scores[0]

    # ── Signal 2: keyword token overlap ──────────────────────────────────────
    def keyword_overlap(row):
        if not query_tokens:
            return 0.0
        title_tokens = set(str(row["clean_title"]).split())
        return len(query_tokens & title_tokens) / len(query_tokens)

    # ── Signal 3: model number / spec exact match boost ───────────────────────
    def model_boost(row):
        if not query_models:
            return 0.0
        row_text   = str(row.get("crawling title", "")) + " " + str(row.get("url_slug", ""))
        row_models = extract_model_numbers(row_text)
        matched    = query_models & row_models
        if not matched:
            return 0.0
        return min(0.20, len(matched) * 0.08)  # each matched code adds 0.08, capped at 0.20

    results["_keyword"] = results.apply(keyword_overlap, axis=1)
    results["_model"]   = results.apply(model_boost,     axis=1)

    # ── Combined score ────────────────────────────────────────────────────────
    results["Similarity Score"] = (
        results["_semantic"] * 0.60
        + results["_keyword"] * 0.20
        + results["_model"]
    )
    # Extra keyword weight for partial queries (brand + model number only)
    if is_partial:
        results["Similarity Score"] += results["_keyword"] * 0.10

    results = results.sort_values("Similarity Score", ascending=False).head(5)
    results = results.drop(columns=["_semantic", "_keyword", "_model"], errors="ignore")

    results["Match Status"] = results["Similarity Score"].apply(
        lambda s: "✅ Match Found" if s >= threshold else "❌ No Match"
    )

    # Store filter metadata as DataFrame attrs (accessible without extra columns)
    results.attrs["domain_filtered"] = domain_filter is not None
    results.attrs["domain_used"]     = domain_filter or ""

    return results


def compare_two_products(
    title_a: str,
    title_b: str,
    model: SentenceTransformer,
    threshold: float = MATCH_THRESHOLD,
) -> dict:
    """
    Directly compare two product titles and return a similarity score.

    Used for URL-to-URL direct comparison mode (when both source and
    target URLs are provided and titles are extracted from both).

    Args:
        title_a:   First product title (source)
        title_b:   Second product title (target)
        model:     Loaded SentenceTransformer
        threshold: Score threshold to determine is_match (default MATCH_THRESHOLD)

    Returns:
        dict with keys:
            score    (float) — cosine similarity 0.0–1.0
            is_match (bool)  — True if score >= threshold
            elapsed  (float) — time taken in seconds

    Example (notebook):
        result = compare_two_products(
            "Samsung Galaxy S24 256GB Blue",
            "سامسونج جالكسي S24 256 جيجا أزرق",
            model
        )
        print(f"Score: {result['score']:.4f}  Match: {result['is_match']}")
    """
    start = time.time()
    emb_a = model.encode([clean_text(title_a)], normalize_embeddings=True).astype("float32")
    emb_b = model.encode([clean_text(title_b)], normalize_embeddings=True).astype("float32")
    score = float(np.dot(emb_a[0], emb_b[0]))

    return {
        "score":    round(score, 4),
        "is_match": score >= threshold,
        "elapsed":  round(time.time() - start, 4),
    }


# ─────────────────────────────────────────────────────────────────────────────
# URL SCRAPING
# ─────────────────────────────────────────────────────────────────────────────

def extract_title_from_url(url: str) -> dict:
    """
    Scrape a product page URL and extract the product title.

    Fallback chain:
        1. og:title meta tag   — most reliable, works on all major sites
        2. Site-specific CSS   — SITE_SELECTORS dict above
        3. <title> tag         — strips site name suffix (| Noon, - Amazon etc.)

    Handles Arabic, Turkish, English page encodings automatically.

    Args:
        url: Full product page URL starting with http/https

    Returns:
        dict with keys:
            title   (str | None)
            source  (str | None) — which method extracted it
            success (bool)
            error   (str | None)

    Example (notebook):
        result = extract_title_from_url("https://www.noon.com/saudi-en/some-product/")
        if result["success"]:
            print(result["title"])
        else:
            print(result["error"])
    """
    if not url or not str(url).strip().startswith("http"):
        return {"title": None, "source": None, "success": False,
                "error": "No valid URL provided (must start with http)"}

    try:
        resp          = requests.get(
            url.strip(), headers=SCRAPE_HEADERS, timeout=10, allow_redirects=True
        )
        resp.encoding = resp.apparent_encoding  # auto-detect Arabic/UTF-8
        soup          = BeautifulSoup(resp.text, "html.parser")

        # 1. Open Graph title — most reliable cross-site
        og = soup.find("meta", property="og:title")
        if og and og.get("content", "").strip():
            return {"title": og["content"].strip(), "source": "og:title",
                    "success": True, "error": None}

        # 2. Site-specific CSS selector
        domain       = re.sub(r"https?://(www\.)?", "", url).split("/")[0]
        matched_site = next((k for k in SITE_SELECTORS if k in domain), "default")
        for selector in SITE_SELECTORS[matched_site]:
            el = soup.select_one(selector)
            if el and el.get_text(strip=True):
                return {"title": el.get_text(strip=True),
                        "source": f"CSS:{selector}", "success": True, "error": None}

        # 3. <title> tag — strip common " | SiteName" or " - SiteName" suffixes
        if soup.title and soup.title.string:
            raw   = soup.title.string.strip()
            title = re.split(r"\s[\|\-–]\s", raw)[0].strip()
            if title:
                return {"title": title, "source": "page_title",
                        "success": True, "error": None}

        return {"title": None, "source": None, "success": False,
                "error": "Product title not found on page"}

    except requests.exceptions.Timeout:
        return {"title": None, "source": None, "success": False,
                "error": "Request timed out (10s) — site may be geo-blocked"}
    except requests.exceptions.ConnectionError:
        return {"title": None, "source": None, "success": False,
                "error": "Connection failed — site may be geo-blocked or offline"}
    except Exception as e:
        return {"title": None, "source": None, "success": False, "error": str(e)}


# ─────────────────────────────────────────────────────────────────────────────
# BATCH PROCESSING
# ─────────────────────────────────────────────────────────────────────────────

def run_batch(
    input_df: pd.DataFrame,
    model: SentenceTransformer,
    index: faiss.IndexFlatIP,
    df: pd.DataFrame,
    batch_threshold: float = MATCH_THRESHOLD,
    global_target: str = "",
    progress_callback=None,
) -> pd.DataFrame:
    """
    Process every row in input_df and return a results DataFrame.

    Target resolution priority (per row):
        1. target_site column value
        2. url_1 column (Excel url.1 field)
        3. global_target argument
        4. No filter — searches all sites

    Args:
        input_df:          DataFrame with columns: product_title, source_url,
                           target_site (all optional but at least one needed)
        model:             Loaded SentenceTransformer
        index:             FAISS index (from build_index)
        df:                Reference product DataFrame (from load_data)
        batch_threshold:   Minimum score to label as "Match Found"
        global_target:     Fallback domain/code if row has no target
        progress_callback: Optional callable(current: int, total: int, msg: str)
                           — used by Streamlit progress bar or notebook print

    Returns:
        pd.DataFrame with one row per input product:
            Row, Source Product, Source URL, Target Website,
            Matched Product, Target Link, Similarity Score,
            Match Status, AHT (seconds)

    Example (notebook):
        input_df = pd.DataFrame({
            "product_title": ["Samsung Galaxy S24", "Xiaomi 14T 256GB"],
            "target_site":   ["noon.com", "TY"]
        })
        results = run_batch(input_df, model, index, df)
        results[["Source Product", "Matched Product", "Similarity Score", "AHT (seconds)"]]
    """
    results_rows = []
    total        = len(input_df)

    for i, row in input_df.iterrows():
        title   = str(row.get("product_title", "")).strip()
        src_url = str(row.get("source_url",    "")).strip()
        url1    = str(row.get("url_1", row.get("url.1", ""))).strip()

        # Target resolution: row → url.1 → global fallback
        raw_target = str(row.get("target_site", "")).strip()
        if not raw_target or raw_target in ("nan", "N/A", ""):
            raw_target = url1
        if not raw_target or raw_target in ("nan", "N/A", ""):
            raw_target = global_target

        resolved_domain = resolve_target_domain(raw_target)

        # Auto-extract title from URL if no title column provided
        if (not title or title in ("nan", "")) and src_url.startswith("http"):
            ext   = extract_title_from_url(src_url)
            title = ext["title"] if ext["success"] else ""

        # Skip row if we still have no title to search with
        if not title or title in ("nan", ""):
            results_rows.append({
                "Row":              i + 1,
                "Source Product":   "⚠️ No title found",
                "Source URL":       src_url or "N/A",
                "Target Website":   resolved_domain or raw_target or "N/A",
                "Matched Product":  "N/A",
                "Target Link":      "N/A",
                "Similarity Score": 0.0,
                "Match Status":     "❌ Skipped — no title",
                "AHT (seconds)":    0.0,
            })
            if progress_callback:
                progress_callback(i + 1, total, f"Skipped row {i + 1} — no title")
            continue

        # Run the match
        t0 = time.time()
        match_results = find_match(
            title, model, index, df,
            threshold=batch_threshold,
            target_filter=resolved_domain if resolved_domain else None,
        )
        elapsed = time.time() - t0

        # Exclude self-match: if the source URL is in the result set, remove it
        if src_url.startswith("http"):
            non_self = match_results[
                match_results["Link"].astype(str).str.strip() != src_url.strip()
            ]
            if not non_self.empty:
                match_results = non_self.reset_index(drop=True)

        best   = match_results.iloc[0]
        score  = float(best["Similarity Score"])
        status = "Match Found" if score >= batch_threshold else "No Match"

        results_rows.append({
            "Row":              i + 1,
            "Source Product":   title,
            "Source URL":       src_url if src_url not in ("nan", "") else "N/A",
            "Target Website":   resolved_domain or raw_target or "N/A",
            "Matched Product":  best["crawling title"],
            "Target Link":      best["Link"],
            "Similarity Score": round(score, 4),
            "Match Status":     status,
            "AHT (seconds)":    round(elapsed, 3),
        })

        if progress_callback:
            progress_callback(i + 1, total, f"Processed: {title[:50]}...")

    return pd.DataFrame(results_rows)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def save_result_with_aht(
    source_title: str,
    source_url: str,
    target_site: str,
    match_title: str,
    target_link: str,
    score: float,
    aht_seconds: float,
    filepath: str = RESULTS_FILE,
) -> pd.DataFrame:
    """
    Append a single search result to the persistent Excel results log.

    Creates the file if it doesn't exist, appends to it if it does.

    Returns:
        Updated DataFrame (all rows including new one)
    """
    status = "Match Found" if score >= MATCH_THRESHOLD else "No Match"
    new_row = {
        "Timestamp":        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Source Product":   source_title,
        "Source URL":       source_url,
        "Target Website":   target_site,
        "Matched Product":  match_title,
        "Target Link":      target_link,
        "Similarity Score": round(score, 4),
        "Match Status":     status,
        "AHT (seconds)":    round(aht_seconds, 2),
    }

    if os.path.exists(filepath):
        existing = pd.read_excel(filepath)
        df_out   = pd.concat([existing, pd.DataFrame([new_row])], ignore_index=True)
    else:
        df_out = pd.DataFrame([new_row])

    _write_excel_with_links(df_out, filepath)
    return df_out


def generate_batch_excel(result_df: pd.DataFrame) -> bytes:
    """
    Serialize batch results to Excel bytes for download.

    Produces a two-sheet workbook:
        Sheet 1 "Batch Results" — full results table with clickable links
        Sheet 2 "Summary"       — match rate + AHT stats

    Returns:
        bytes — pass directly to st.download_button(data=...) or write to file

    Example (notebook):
        excel_bytes = generate_batch_excel(results)
        with open("output.xlsx", "wb") as f:
            f.write(excel_bytes)
    """
    buffer   = io.BytesIO()
    wb       = openpyxl.Workbook()
    ws       = wb.active
    ws.title = "Batch Results"

    headers    = list(result_df.columns)
    link_cols  = {"Source URL", "Target Link"}
    col_widths = {
        "Row": 6, "Source Product": 35, "Source URL": 45,
        "Target Website": 25, "Matched Product": 35,
        "Target Link": 45, "Similarity Score": 16,
        "Match Status": 16, "AHT (seconds)": 14,
    }

    _apply_header_style(ws, headers)

    for row_idx, row in result_df.iterrows():
        _write_data_row(ws, row_idx + 2, row_idx, headers, row, link_cols)

    for col_idx, col_name in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)
    ws.freeze_panes = "A2"

    _write_summary_sheet(wb.create_sheet("Summary"), result_df)

    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL HELPERS (internal — not exported)
# ─────────────────────────────────────────────────────────────────────────────

def _get_excel_styles() -> dict:
    """Return a dict of reusable openpyxl style objects."""
    thin = Side(border_style="thin", color="CCCCCC")
    return {
        "header_fill": PatternFill("solid", fgColor="1F4E79"),
        "header_font": Font(name="Arial", bold=True, color="FFFFFF", size=11),
        "center":      Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border":      Border(left=thin, right=thin, top=thin, bottom=thin),
    }


def _apply_header_style(ws, headers: list):
    """Write styled header row to a worksheet."""
    s = _get_excel_styles()
    for col_idx, header in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = s["header_fill"]
        cell.font      = s["header_font"]
        cell.alignment = s["center"]
        cell.border    = s["border"]
    ws.row_dimensions[1].height = 30


def _write_data_row(ws, excel_row: int, row_idx: int, headers: list, row, link_cols: set):
    """Write a single styled data row to a worksheet."""
    s = _get_excel_styles()
    for col_idx, col_name in enumerate(headers, start=1):
        val  = row[col_name]
        cell = ws.cell(row=excel_row, column=col_idx)
        cell.border    = s["border"]
        cell.alignment = Alignment(vertical="center")
        cell.font      = Font(name="Arial", size=10)

        if col_name in link_cols and str(val).startswith("http"):
            cell.value     = str(val)
            cell.hyperlink = str(val)
            cell.font      = Font(name="Arial", size=10, color="0563C1", underline="single")
        elif col_name == "Match Status":
            cell.value     = val
            cell.fill      = PatternFill("solid", fgColor="C6EFCE" if val == "Match Found" else "FFCCCC")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_name == "Similarity Score":
            cell.value         = val
            cell.number_format = "0.0000"
            cell.alignment     = Alignment(horizontal="center", vertical="center")
        else:
            cell.value = val

    # Alternating row stripe
    if row_idx % 2 == 0:
        for ci in range(1, len(headers) + 1):
            c = ws.cell(row=excel_row, column=ci)
            if c.fill.fgColor.rgb == "00000000":
                c.fill = PatternFill("solid", fgColor="F2F7FC")


def _write_summary_sheet(ws2, result_df: pd.DataFrame):
    """Write a summary stats sheet to an already-created worksheet."""
    ws2["A1"]      = "Batch Processing Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")

    total   = len(result_df)
    matched = len(result_df[result_df["Match Status"] == "Match Found"])

    rows = [
        ("Total Products",       total),
        ("Matched",              matched),
        ("Not Matched",          total - matched),
        ("Match Rate",           f"{round(matched / total * 100, 1)}%" if total else "0%"),
        ("Avg AHT (seconds)",    round(result_df["AHT (seconds)"].mean(), 3) if total else 0),
        ("Total Time (seconds)", round(result_df["AHT (seconds)"].sum(),  2) if total else 0),
    ]
    for r, (label, value) in enumerate(rows, start=3):
        ws2.cell(r, 1, label).font = Font(name="Arial", bold=True, size=11)
        ws2.cell(r, 2, value).font = Font(name="Arial", size=11)

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 18


def _write_excel_with_links(df_out: pd.DataFrame, filepath: str):
    """Write a DataFrame to an Excel file with clickable links and an AHT summary sheet."""
    wb       = openpyxl.Workbook()
    ws       = wb.active
    ws.title = "Matching Results"

    headers   = list(df_out.columns)
    link_cols = {"Source URL", "Target Link"}
    col_widths = {
        "Timestamp": 20, "Source Product": 35, "Source URL": 45,
        "Target Website": 25, "Matched Product": 35, "Target Link": 45,
        "Similarity Score": 16, "Match Status": 16, "AHT (seconds)": 16,
    }

    _apply_header_style(ws, headers)
    for row_idx, row in df_out.iterrows():
        _write_data_row(ws, row_idx + 2, row_idx, headers, row, link_cols)
    for col_idx, col_name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)
    ws.freeze_panes = "A2"

    # AHT summary tab
    ws2       = wb.create_sheet("AHT Summary")
    ws2["A1"] = "AHT Analytics"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")

    has_aht     = "AHT (seconds)" in df_out.columns and len(df_out) > 0
    has_status  = "Match Status"  in df_out.columns and len(df_out) > 0
    match_count = len(df_out[df_out["Match Status"] == "Match Found"]) if has_status else 0

    aht_rows = [
        ("Total Searches",    len(df_out)),
        ("Avg AHT (seconds)", round(df_out["AHT (seconds)"].mean(), 2) if has_aht else 0),
        ("Min AHT (seconds)", round(df_out["AHT (seconds)"].min(),  2) if has_aht else 0),
        ("Max AHT (seconds)", round(df_out["AHT (seconds)"].max(),  2) if has_aht else 0),
        ("Match Rate",        f"{round(match_count / len(df_out) * 100, 1)}%" if has_status else "0%"),
    ]
    for r, (label, value) in enumerate(aht_rows, start=3):
        ws2.cell(r, 1, label).font = Font(name="Arial", bold=True, size=11)
        ws2.cell(r, 2, value).font = Font(name="Arial", size=11)

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 20
    wb.save(filepath)


# ─────────────────────────────────────────────────────────────────────────────
# TERMINAL SELF-TEST  —  python utils/matcher.py
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":

    print("=" * 62)
    print("  matcher.py — Terminal Self-Test")
    print("=" * 62)

    all_passed = True

    # ── Step 1: Data loading ──────────────────────────────────────────────────
    print("\n[1/6] Loading data...")
    try:
        _df = load_data()
        print(f"  ✅ Loaded {len(_df)} products")
        print(f"     Columns : {list(_df.columns)}")
        assert "clean_title"    in _df.columns, "clean_title missing"
        assert "enriched_title" in _df.columns, "enriched_title missing"
        assert "url_slug"       in _df.columns, "url_slug missing"
        assert len(_df) > 0,                    "DataFrame is empty"
    except FileNotFoundError as e:
        print(f"  ⚠️  Data file not found — skipping data-dependent steps")
        print(f"     ({e})")
        print(f"\n  To run the full self-test, place your Excel file at:")
        print(f"     {DEFAULT_DATA_PATH}")
        _df = None

    # ── Step 2: Text cleaning ─────────────────────────────────────────────────
    print("\n[2/6] Testing clean_text()...")
    try:
        _cases = [
            ("Samsung Galaxy 256GB",     "samsung galaxy 256 gb"),
            ("سامسونج S24",              "سامسونج s24"),
            ("HP!!  Laptop--15s  ",      "hp laptop 15s"),
        ]
        for raw, expected in _cases:
            result = clean_text(raw)
            assert expected in result or result.startswith(expected.split()[0]), \
                f"clean_text({raw!r}) = {result!r}, expected to contain {expected!r}"
        print(f"  ✅ clean_text() — {len(_cases)} test cases passed")
    except AssertionError as e:
        print(f"  ❌ {e}")
        all_passed = False

    # ── Step 3: Domain helpers ────────────────────────────────────────────────
    print("\n[3/6] Testing domain helpers...")
    try:
        _domain_tests = [
            ("TY",                          "trendyol.com"),
            ("NN_KSA",                      "noon.com"),
            ("www.noon.com",                "noon.com"),
            ("https://extra.com/p/123",     "extra.com"),
            ("trendyol.com",                "trendyol.com"),
            ("",                            ""),
        ]
        for inp, expected in _domain_tests:
            result = resolve_target_domain(inp)
            assert result == expected, \
                f"resolve_target_domain({inp!r}) = {result!r}, expected {expected!r}"
        print(f"  ✅ resolve_target_domain() — {len(_domain_tests)} test cases passed")
    except AssertionError as e:
        print(f"  ❌ {e}")
        all_passed = False

    # ── Step 4: Model number extractor ───────────────────────────────────────
    print("\n[4/6] Testing extract_model_numbers()...")
    try:
        _model_tests = [
            ("Samsung 256GB NFK37211", {"NFK37211", "256GB"}),
            ("HP 2S9L1AA laptop",      {"2S9L1AA"}),
            ("plain product name",     set()),
        ]
        for text, expected_subset in _model_tests:
            result = extract_model_numbers(text)
            for code in expected_subset:
                assert code in result, \
                    f"extract_model_numbers({text!r}) missing {code!r}, got {result}"
        print(f"  ✅ extract_model_numbers() — {len(_model_tests)} test cases passed")
    except AssertionError as e:
        print(f"  ❌ {e}")
        all_passed = False

    # ── Step 5: Model loading ─────────────────────────────────────────────────
    print("\n[5/6] Loading SentenceTransformer model...")
    try:
        _t0    = time.time()
        _model = load_model()
        _elapsed = time.time() - _t0
        _dim   = _model.get_sentence_embedding_dimension()
        assert _dim == 384, f"Expected embedding dim 384, got {_dim}"
        print(f"  ✅ Model loaded in {_elapsed:.1f}s  (dim={_dim})")
    except Exception as e:
        print(f"  ❌ Model load failed: {e}")
        all_passed = False
        _model = None

    # ── Step 6: Full matching pipeline ───────────────────────────────────────
    print("\n[6/6] Running end-to-end match test...")
    if _df is not None and _model is not None:
        try:
            _t0    = time.time()
            _index = build_index(_df, _model, show_progress=False)
            _build_time = time.time() - _t0
            assert _index.ntotal == len(_df), \
                f"Index has {_index.ntotal} vectors but df has {len(_df)} rows"
            print(f"  ✅ FAISS index built in {_build_time:.1f}s  ({_index.ntotal} vectors)")

            _query   = str(_df["crawling title"].iloc[0])  # use first product as test query
            _t0      = time.time()
            _results = find_match(_query, _model, _index, _df, threshold=0.70)
            _match_time = time.time() - _t0
            assert len(_results) > 0,          "find_match returned empty DataFrame"
            assert "Similarity Score" in _results.columns, "Similarity Score column missing"
            assert "Match Status"     in _results.columns, "Match Status column missing"
            _top = _results.iloc[0]
            print(f"  ✅ find_match() ran in {_match_time:.3f}s")
            print(f"     Query       : {_query[:60]}")
            print(f"     Top result  : {str(_top['crawling title'])[:60]}")
            print(f"     Score       : {_top['Similarity Score']:.4f}")
            print(f"     Status      : {_top['Match Status']}")

            # compare_two_products test
            _comp = compare_two_products(_query, str(_top["crawling title"]), _model)
            assert "score"    in _comp, "score key missing from compare_two_products result"
            assert "is_match" in _comp, "is_match key missing"
            assert "elapsed"  in _comp, "elapsed key missing"
            print(f"  ✅ compare_two_products() → score={_comp['score']:.4f}")

        except Exception as e:
            print(f"  ❌ Pipeline test failed: {e}")
            all_passed = False
    else:
        print("  ⏭️  Skipped — data or model not available")

    # ── Summary ───────────────────────────────────────────────────────────────
    print()
    print("=" * 62)
    if all_passed:
        print("  ✅ All checks passed — matcher.py is working correctly!")
    else:
        print("  ❌ Some checks failed — see details above")
    print("=" * 62)

    # ── Interactive Quick-Test Mode ────────────────────────────────────────────
    # Runs ONLY when data + model are available and all checks passed.
    # Skip with Ctrl+C at any time.
    if all_passed and _df is not None and _model is not None:
        print()
        print("=" * 62)
        print("  🧪 INTERACTIVE QUICK-TEST  (Ctrl+C to exit)")
        print("=" * 62)
        print("  Test single product OR batch file from the command line.")
        print()
        print("  Options:")
        print("    [1] Single product title search")
        print("    [2] Batch test from CSV/Excel file")
        print("    [3] Exit")
        print()

        _index = build_index(_df, _model, show_progress=False)

        while True:
            try:
                choice = input("  Choose option (1/2/3): ").strip()

                # ── Option 1: Single query ────────────────────────────────────
                if choice == "1":
                    query = input("\n  Enter product title (any language): ").strip()
                    if not query:
                        print("  ⚠️  Empty input — skipping.\n")
                        continue

                    target = input(
                        "  Target site/code (optional, press Enter to skip): "
                    ).strip() or None

                    thresh_raw = input(
                        f"  Threshold (default {MATCH_THRESHOLD}, press Enter to keep): "
                    ).strip()
                    try:
                        thresh = float(thresh_raw) if thresh_raw else MATCH_THRESHOLD
                    except ValueError:
                        thresh = MATCH_THRESHOLD

                    print()
                    t0       = time.time()
                    results  = find_match(query, _model, _index, _df,
                                          threshold=thresh, target_filter=target)
                    elapsed  = time.time() - t0

                    print(f"  ⏱  Completed in {elapsed:.3f}s")
                    print(f"  🔍 Query       : {query}")
                    if target:
                        print(f"  🌐 Target      : {target}  →  {results.attrs.get('domain_used', '')}")
                    print(f"  📊 Top 5 results (threshold = {thresh}):")
                    print()

                    # Pretty-print results table
                    print(f"  {'#':<3} {'Score':<8} {'Status':<18} {'Title':<45} {'Link'}")
                    print("  " + "-" * 110)
                    for rank, (_, row) in enumerate(results.iterrows(), start=1):
                        title_short = str(row['crawling title'])[:43]
                        link_short  = str(row['Link'])[:55]
                        score_val   = row['Similarity Score']
                        status_val  = row['Match Status']
                        print(f"  {rank:<3} {score_val:<8.4f} {status_val:<18} {title_short:<45} {link_short}")
                    print()

                # ── Option 2: Batch file ──────────────────────────────────────
                elif choice == "2":
                    file_path = input(
                        "\n  Enter path to CSV or Excel file: "
                    ).strip().strip('"').strip("'")

                    if not file_path or not Path(file_path).exists():
                        print(f"  ❌ File not found: {file_path}\n")
                        continue

                    print(f"  📂 Loading: {file_path} ...")
                    try:
                        if file_path.lower().endswith(".csv"):
                            raw_df = pd.read_csv(file_path)
                        else:
                            raw_df = pd.read_excel(file_path)
                    except Exception as e:
                        print(f"  ❌ Could not read file: {e}\n")
                        continue

                    print(f"  ✅ Loaded {len(raw_df)} rows")
                    print(f"     Columns: {list(raw_df.columns)}")
                    print()

                    # Auto-detect columns
                    cols_lower = {c.lower().replace(" ", "_"): c for c in raw_df.columns}
                    title_col  = next((cols_lower[k] for k in
                                       ["crawling_title", "title", "product_title", "product", "name"]
                                       if k in cols_lower), None)
                    url_col    = next((cols_lower[k] for k in
                                       ["url", "source_url", "link", "source"]
                                       if k in cols_lower), None)
                    target_col = next((cols_lower[k] for k in
                                       ["target_retailer_code", "target_site", "target", "url.1", "url_1"]
                                       if k in cols_lower), None)

                    print(f"  Auto-detected columns:")
                    print(f"    Title  : {title_col or '❌ not found'}")
                    print(f"    URL    : {url_col   or '(not found — optional)'}")
                    print(f"    Target : {target_col or '(not found — optional)'}")
                    print()

                    if title_col is None and url_col is None:
                        print("  ❌ Could not find a title or URL column. "
                              "Please rename your column to 'title' or 'crawling title'.\n")
                        continue

                    # Build input_df for run_batch
                    input_df = pd.DataFrame()
                    if title_col:
                        input_df["product_title"] = raw_df[title_col].astype(str).str.strip()
                    if url_col:
                        input_df["source_url"] = raw_df[url_col].astype(str).str.strip()
                    if target_col:
                        input_df["target_site"] = raw_df[target_col].astype(str).str.strip()

                    global_target = input(
                        "  Global target site/code (optional, press Enter to skip): "
                    ).strip()

                    thresh_raw = input(
                        f"  Threshold (default {MATCH_THRESHOLD}, press Enter to keep): "
                    ).strip()
                    try:
                        thresh = float(thresh_raw) if thresh_raw else MATCH_THRESHOLD
                    except ValueError:
                        thresh = MATCH_THRESHOLD

                    print(f"\n  ⚙️  Processing {len(input_df)} products...\n")

                    def _cli_progress(current, total, msg):
                        bar_len   = 30
                        filled    = int(bar_len * current / total)
                        bar       = "█" * filled + "░" * (bar_len - filled)
                        pct       = current / total * 100
                        print(f"\r  [{bar}] {pct:.0f}%  {msg[:40]:<40}", end="", flush=True)

                    batch_results = run_batch(
                        input_df, _model, _index, _df,
                        batch_threshold=thresh,
                        global_target=global_target,
                        progress_callback=_cli_progress,
                    )
                    print()  # newline after progress bar

                    # Summary stats
                    total_r   = len(batch_results)
                    matched_r = len(batch_results[batch_results["Match Status"] == "Match Found"])
                    avg_aht   = batch_results["AHT (seconds)"].mean()

                    print()
                    print(f"  ✅ Done!")
                    print(f"  📊 Results:")
                    print(f"     Total     : {total_r}")
                    print(f"     Matched   : {matched_r}  ({matched_r/total_r*100:.1f}%)")
                    print(f"     No Match  : {total_r - matched_r}")
                    print(f"     Avg AHT   : {avg_aht:.4f}s per product")
                    print()

                    # Pretty print results
                    print(f"  {'Row':<5} {'Score':<8} {'Status':<16} {'Source (truncated)':<35} {'Matched (truncated)':<35}")
                    print("  " + "-" * 103)
                    for _, row in batch_results.iterrows():
                        src   = str(row["Source Product"])[:33]
                        match = str(row["Matched Product"])[:33]
                        print(f"  {int(row['Row']):<5} {row['Similarity Score']:<8.4f} "
                              f"{str(row['Match Status']):<16} {src:<35} {match:<35}")
                    print()

                    # Offer to save output
                    save_path = input(
                        "  Save results to Excel? Enter filename (or press Enter to skip): "
                    ).strip()
                    if save_path:
                        if not save_path.endswith(".xlsx"):
                            save_path += ".xlsx"
                        excel_bytes = generate_batch_excel(batch_results)
                        with open(save_path, "wb") as f:
                            f.write(excel_bytes)
                        print(f"  ✅ Saved to: {save_path}\n")
                    else:
                        print()

                # ── Option 3: Exit ────────────────────────────────────────────
                elif choice == "3":
                    print("  👋 Exiting quick-test. Goodbye!\n")
                    break
                else:
                    print("  ⚠️  Invalid choice. Enter 1, 2, or 3.\n")

            except KeyboardInterrupt:
                print("\n\n  👋 Interrupted. Goodbye!\n")
                break
            except Exception as e:
                print(f"\n  ❌ Unexpected error: {e}\n")
